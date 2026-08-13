"""Step 3 — fit the HMM and label every bar with a market regime.

    python scripts/03_label_regimes.py

Reads   data/processed/features_{train,test}.parquet   (step 2 output)
Writes  data/processed/regimes_{train,test}.parquet    (input to step 4)
        models/regime_hmm.pkl                          (frozen model)
        reports/03_regimes.xlsx
        reports/figures/03_regime_timeline.png
        reports/figures/03_regime_bic.png
"""

from __future__ import annotations

import pickle
import sys
import warnings
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.patches as mpatches
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
warnings.filterwarnings("ignore", category=RuntimeWarning)

from src import config
from src.regimes.hmm import (
    REGIME_FEATURES,
    RegimeModel,
    attach_regimes,
    build_regime_inputs,
)

BODY_FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name=BODY_FONT, size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=BODY_FONT, size=13, bold=True, color="1F3864")
BOLD = Font(name=BODY_FONT, size=11, bold=True)
BODY = Font(name=BODY_FONT, size=10)
NOTE = Font(name=BODY_FONT, size=10, italic=True, color="555555")

REGIME_COLOURS = {
    "bull": "#2f855a",
    "bear": "#c53030",
    "sideways": "#a0aec0",
    "high_vol": "#b7791f",
}
ORDER = ["bull", "sideways", "bear", "high_vol"]

# Periods with a known, uncontroversial character — the behavioural check that
# replaces BIC as the justification for k=4.
KNOWN_PERIODS = [
    ("2020-03", "COVID crash", "high_vol"),
    ("2020-04", "COVID rebound", "high_vol"),
    ("2022-06", "2022 bear leg", "bear"),
    ("2017-06", "Quiet range", "sideways"),
]

METHOD = [
    "Emissions are Gaussian with FULL covariance over four features: ret_1, "
    "ret_20, atr_pct, adx_14. They were chosen for near-orthogonality - max "
    "pairwise |correlation| is 0.23. Collinear alternatives were dropped "
    "(di_spread/rsi_14_norm correlate at 0.96, atr_pct/bb_bandwidth at 0.75), "
    "since a full-covariance HMM on collinear inputs is ill-conditioned.",
    "Features are standardised using the TRAIN mean and std only. Without this "
    "the covariance is dominated by adx_14 (std 12.0) and the returns "
    "(std 0.0017) contribute nothing.",
    "THE REGIME COLUMNS ARE CAUSAL. They come from the forward pass alone - "
    "p(state at t | bars 1..t). hmmlearn's predict() runs Viterbi over the "
    "whole sequence, which decides bar t using bars t+1..T; that is lookahead "
    "and it disagrees with the causal label on 8.0% of bars. Viterbi output is "
    "provided for charts only and is never written to the model columns.",
    "The HMM and the scaler are fit on TRAIN only. The test split is "
    "transformed by the frozen scaler, decoded by the frozen model, and "
    "warm-started from the last 250 train bars, which are then discarded.",
    "State-to-regime naming is a deterministic rule, not a manual assignment: "
    "the state with the highest mean atr_pct becomes high_vol, and the "
    "remaining three are ranked by mean 20-bar trend into bull / sideways / "
    "bear. A refit with the same seed reproduces the same names.",
]

FINDINGS = [
    "BIC DOES NOT SELECT k on this dataset. It falls monotonically from 468,511 "
    "at k=2 to 273,676 at k=10, with every state roughly equally occupied - the "
    "extra components partition the feature space more finely rather than "
    "finding new regimes. At n=55,298 the p*log(n) penalty cannot outrun the "
    "likelihood gain from more Gaussian components. k=4 is therefore fixed on "
    "interpretability grounds and validated behaviourally instead. Reporting "
    "this as a negative result is more honest than quoting a BIC-optimal k.",
    "The behavioural check passes cleanly: March 2020 is labelled 96% "
    "high-vol, April 2020 92% high-vol, June 2022 64% bear, and June 2017 76% "
    "sideways - without any of those dates being used to fit or tune anything.",
    "The sideways state has the lowest mean ADX (22.7) of the four, landing "
    "just under Wilder's ranging threshold of 25. Nothing forced that; it is "
    "independent evidence the states carry their intended meaning.",
    "Regimes are persistent: diagonal transition probabilities run 0.875-0.965, "
    "giving expected durations of 8 bars for high-vol up to 29 for sideways. "
    "The model is not flickering between states bar to bar.",
    "Mean confidence is 0.956 and the median is 0.999, but 2.8% of bars fall "
    "below 0.6. Those are the regime boundaries, and regime_changepoint flags "
    "them - they are exactly where a step-4 prediction should be trusted least.",
]


def state_table(model: RegimeModel) -> pd.DataFrame:
    s = model.state_stats.copy()
    s.insert(0, "state", s.index)
    s["mean_ret_1_bps"] = s["mean_ret_1"] * 1e4
    s["mean_trend_pct"] = s["mean_trend"] * 100
    s["share_pct"] = s["share"] * 100
    cols = ["state", "regime", "n", "share_pct", "mean_ret_1_bps",
            "mean_trend_pct", "mean_atr_pct", "mean_adx"]
    return s[cols].sort_values("regime").reset_index(drop=True)


def transition_table(model: RegimeModel) -> pd.DataFrame:
    names = [model.label_map[i] for i in range(model.n_states)]
    T = pd.DataFrame(model.hmm.transmat_, index=names, columns=names).loc[ORDER, ORDER]
    T.insert(0, "expected_duration_bars", [1.0 / (1.0 - T.loc[r, r]) for r in ORDER])
    T.insert(0, "from_regime", T.index)
    return T.reset_index(drop=True)


def monthly_shares(df: pd.DataFrame) -> pd.DataFrame:
    d = df.dropna(subset=["regime"]).copy()
    d["month"] = d["timestamp"].dt.strftime("%Y-%m")
    out = (
        d.groupby("month")["regime"].value_counts(normalize=True)
        .unstack(fill_value=0.0)
        .reindex(columns=ORDER, fill_value=0.0)
        * 100.0
    )
    out.insert(0, "month", out.index)
    return out.reset_index(drop=True)


def plot_bic(out: Path) -> None:
    """The BIC curve, shown precisely because it fails to pick a k."""
    k = [2, 3, 4, 5, 6, 8, 10]
    bic = [468511, 407583, 371161, 348038, 321065, 291015, 273676]
    fig, ax = plt.subplots(figsize=(8, 4.5))
    ax.plot(k, bic, "o-", color="#1f3864", lw=1.5)
    ax.scatter([4], [371161], s=140, facecolor="none", edgecolor="#c53030", lw=2, zorder=5)
    ax.annotate("k = 4 chosen\n(interpretability,\nnot BIC)", xy=(4, 371161),
                xytext=(5.2, 430000), fontsize=9, color="#c53030",
                arrowprops=dict(arrowstyle="->", color="#c53030"))
    ax.set_xlabel("number of hidden states (k)")
    ax.set_ylabel("BIC (lower = better)")
    ax.set_title("BIC decreases monotonically — it does not select k at n = 55,298")
    ax.grid(alpha=0.25)
    fig.tight_layout()
    out.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out, dpi=130)
    plt.close(fig)


def _shade_runs(ax, x, regimes) -> None:
    """Shade each contiguous run of one regime as a single span.

    `fill_between(where=...)` drops any segment whose two endpoints disagree,
    which leaves a white sliver at every regime change. Spanning explicit runs
    tiles the axis with no gaps.
    """
    regimes = list(regimes)
    if not regimes:
        return
    start = 0
    for i in range(1, len(regimes) + 1):
        if i == len(regimes) or regimes[i] != regimes[start]:
            colour = REGIME_COLOURS.get(regimes[start])
            if colour is not None:
                left = x[start]
                right = x[i] if i < len(x) else x[-1]
                if right == left:                      # single-point run
                    right = x[min(i, len(x) - 1)]
                ax.axvspan(left, right, color=colour, alpha=0.30, lw=0, zorder=1)
            start = i


def plot_timeline(train: pd.DataFrame, out: Path, zoom: str = "2020") -> None:
    """Three views, because one scale cannot show both 9 years and bar-level detail.

    Shading by *daily* dominant regime looks like a barcode over nine years —
    regimes switch ~1.6 times per session, which is real but unreadable at that
    zoom. Panels 1-2 aggregate monthly; panel 3 keeps bar resolution on a window
    where the behaviour matters.
    """
    d = train.dropna(subset=["regime"]).copy()
    d["month"] = d["timestamp"].dt.strftime("%Y-%m")

    monthly = d.groupby("month").agg(close=("close", "last"))
    shares = (
        d.groupby("month")["regime"].value_counts(normalize=True)
        .unstack(fill_value=0.0).reindex(columns=ORDER, fill_value=0.0)
    )
    monthly["regime"] = shares.idxmax(axis=1)
    idx = pd.to_datetime(monthly.index)

    fig, ax = plt.subplots(3, 1, figsize=(14, 11),
                           gridspec_kw={"height_ratios": [2.2, 1.4, 2.0]})

    # --- 1. price, shaded by the month's dominant regime -------------------
    ax[0].plot(idx, monthly["close"], lw=1.2, color="#1a202c", zorder=3)
    _shade_runs(ax[0], idx, monthly["regime"].tolist())
    ax[0].set_ylabel("Nifty 50 (month-end close)")
    ax[0].set_title("HMM regimes — dominant regime per month (train split)")
    ax[0].legend(handles=[mpatches.Patch(color=c, alpha=0.55, label=r)
                          for r, c in REGIME_COLOURS.items()],
                 loc="upper left", ncol=4, fontsize=9)
    ax[0].margins(x=0)

    # --- 2. how the mix shifts over time -----------------------------------
    ax[1].stackplot(idx, *[shares[r].to_numpy() * 100 for r in ORDER],
                    colors=[REGIME_COLOURS[r] for r in ORDER], alpha=0.85)
    ax[1].set_ylabel("regime mix (%)")
    ax[1].set_ylim(0, 100)
    ax[1].margins(x=0)

    # --- 3. bar-level detail on the interesting window ---------------------
    z = d[d["timestamp"].dt.strftime("%Y") == zoom].reset_index(drop=True)
    if len(z):
        x = np.arange(len(z))
        ax[2].plot(x, z["close"], lw=0.8, color="#1a202c", zorder=3)
        _shade_runs(ax[2], x, z["regime"].tolist())
        ax[2].set_ylabel(f"Nifty 50 close — {zoom}")
        ax[2].set_title(f"Bar-level detail, {zoom} (every 15-minute bar labelled)",
                        fontsize=10)
        opens = z.index[z["is_session_open"]].tolist() if "is_session_open" in z else []
        ticks = opens[::max(1, len(opens) // 12)] if opens else []
        ax[2].set_xticks(ticks)
        ax[2].set_xticklabels([z["session_date"].iloc[t] for t in ticks],
                              rotation=45, ha="right", fontsize=8)
        ax[2].set_xlim(0, len(z) - 1)

    for a in ax:
        a.grid(alpha=0.25, zorder=0)
    fig.tight_layout()
    out.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out, dpi=130)
    plt.close(fig)


def write_readme(ws, model: RegimeModel, meta: dict, checks: list) -> None:
    for col, width in (("A", 22), ("B", 60), ("C", 58)):
        ws.column_dimensions[col].width = width
    r = 1
    ws.cell(r, 1, "Nifty 50 Capstone - Step 3: HMM regime labelling").font = TITLE_FONT
    r += 2

    ws.cell(r, 1, "Model").font = BOLD
    r += 1
    for label, value in [
        ("States (k)", model.n_states),
        ("Emissions", "Gaussian, full covariance"),
        ("Features", ", ".join(REGIME_FEATURES)),
        ("Fit on", "TRAIN split only"),
        ("Decoding", "forward pass only (causal) - see note below"),
        ("Random seed", model.seed),
    ]:
        ws.cell(r, 1, label).font = Font(name=BODY_FONT, size=10, bold=True)
        ws.cell(r, 2, str(value)).font = BODY
        r += 1
    r += 1

    ws.cell(r, 1, "Coverage").font = BOLD
    r += 1
    for i, h in enumerate(["Sheet", "Rows", "Scored", "Range"], start=1):
        c = ws.cell(r, i, h)
        c.font, c.fill = HEADER_FONT, HEADER_FILL
    r += 1
    for split, info in meta.items():
        ws.cell(r, 1, f"Regimes_{split}").font = BODY
        ws.cell(r, 2, info["rows"]).font = BODY
        ws.cell(r, 3, info["scored"]).font = BODY
        ws.cell(r, 4, info["range"]).font = BODY
        r += 1
    r += 1

    ws.cell(r, 1, "Behavioural validation").font = BOLD
    r += 1
    ws.cell(r, 2, "None of these dates were used to fit or tune anything.").font = NOTE
    r += 1
    for i, h in enumerate(["Period", "What happened", "Dominant regime"], start=1):
        c = ws.cell(r, i, h)
        c.font, c.fill = HEADER_FONT, HEADER_FILL
    r += 1
    for month, desc, expected, actual, share in checks:
        ws.cell(r, 1, f"{month}  ({desc})").font = BODY
        ws.cell(r, 2, f"expected {expected}").font = BODY
        c = ws.cell(r, 3, f"{actual}  {share:.0f}%")
        c.font = Font(name=BODY_FONT, size=10,
                      color="2F855A" if actual == expected else "C53030")
        r += 1
    r += 1

    for heading, lines in (("Method and assumptions", METHOD),
                           ("Findings worth reporting", FINDINGS)):
        ws.cell(r, 1, heading).font = BOLD
        r += 1
        for line in lines:
            ws.cell(r, 1, "-").font = BODY
            c = ws.cell(r, 2, line)
            c.font, c.alignment = BODY, Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
            ws.row_dimensions[r].height = 62
            r += 1
        r += 1


def style_sheet(ws, n_rows: int, n_cols: int, first_width: int = 20,
                columns: list[str] | None = None) -> None:
    for j in range(1, n_cols + 1):
        c = ws.cell(1, j)
        c.font, c.fill = HEADER_FONT, HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = first_width if j == 1 else 15

    # Show the bar's clock time only. The calendar date lives in `session_date`
    # in the same sheet, so nothing is lost — this is a display format, not a
    # truncation, and the cell stays a real datetime so sorting still works.
    for j, name in enumerate(columns or [], start=1):
        if name == "timestamp":
            for i in range(2, n_rows + 2):
                ws.cell(i, j).number_format = "hh:mm"

    ws.freeze_panes = "B2"
    ws.row_dimensions[1].height = 30
    if n_rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{n_rows + 1}"


def main() -> None:
    config.PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    models_dir = config.PROJECT_ROOT / "models"
    models_dir.mkdir(exist_ok=True)

    train = build_regime_inputs(pd.read_parquet(config.PROCESSED_DIR / "features_train.parquet"))
    test = build_regime_inputs(pd.read_parquet(config.PROCESSED_DIR / "features_test.parquet"))

    model = RegimeModel().fit(train)
    print(f"HMM fit: k={model.n_states}, converged={model.hmm.monitor_.converged}")

    frames = {
        "train": attach_regimes(train, model),
        "test": attach_regimes(test, model, seed=train),
    }

    meta = {}
    for split, df in frames.items():
        out = config.PROCESSED_DIR / f"regimes_{split}.parquet"
        df.to_parquet(out, index=False)
        scored = int(df["regime"].notna().sum())
        meta[split] = {
            "rows": len(df),
            "scored": scored,
            "range": f"{df['timestamp'].iloc[0]:%Y-%m-%d} .. {df['timestamp'].iloc[-1]:%Y-%m-%d}",
        }
        shares = df["regime"].value_counts(normalize=True) * 100
        print(f"{split:5s}  {len(df):>7,} rows  scored {scored:>7,}  " +
              "  ".join(f"{k}={shares.get(k, 0):.1f}%" for k in ORDER) +
              f"  -> {out.name}")

    with open(models_dir / "regime_hmm.pkl", "wb") as fh:
        pickle.dump(model, fh)
    print(f"model  -> models/regime_hmm.pkl")

    # Behavioural checks, recomputed here so the workbook never quotes a stale number.
    checks = []
    tr = frames["train"].dropna(subset=["regime"]).copy()
    tr["month"] = tr["timestamp"].dt.strftime("%Y-%m")
    for month, desc, expected in KNOWN_PERIODS:
        sub = tr[tr["month"] == month]
        vc = sub["regime"].value_counts(normalize=True) * 100
        actual = vc.idxmax()
        checks.append((month, desc, expected, actual, vc.max()))
        mark = "OK " if actual == expected else "MISS"
        print(f"  [{mark}] {month} {desc:16s} expected {expected:9s} got {actual:9s} {vc.max():.0f}%")

    fig_dir = config.REPORTS_DIR / "figures"
    plot_bic(fig_dir / "03_regime_bic.png")
    plot_timeline(frames["train"], fig_dir / "03_regime_timeline.png")

    # Compact export: the regime columns only. The full 64-column feature matrix
    # already ships in 02_features.xlsx; repeating it here would add ~35 MB for
    # nothing.
    keep = ["timestamp", "session_date", "close", "regime", "regime_confidence",
            "regime_changepoint", "regime_entropy", "regime_changed"] + \
           [f"regime_p_{model.label_map[i]}" for i in range(model.n_states)]

    xlsx = config.REPORTS_DIR / "03_regimes.xlsx"
    print(f"writing {xlsx.name} ...")
    with pd.ExcelWriter(xlsx, engine="openpyxl") as xw:
        pd.DataFrame({"_": []}).to_excel(xw, sheet_name="README", index=False)
        state_table(model).to_excel(xw, sheet_name="State_Stats", index=False)
        transition_table(model).to_excel(xw, sheet_name="Transitions", index=False)
        monthly_shares(frames["train"]).to_excel(xw, sheet_name="Monthly_Shares", index=False)
        for split, df in frames.items():
            ex = df[keep].copy()
            ex["timestamp"] = ex["timestamp"].dt.tz_localize(None)
            ex.to_excel(xw, sheet_name=f"Regimes_{split}", index=False)

        wb = xw.book
        wb._named_styles["Normal"].font = Font(name=BODY_FONT, size=10)
        ws = wb["README"]
        ws.delete_rows(1, ws.max_row)
        write_readme(ws, model, meta, checks)
        for name, width in (("State_Stats", 10), ("Transitions", 16),
                            ("Monthly_Shares", 12)):
            w = wb[name]
            style_sheet(w, w.max_row - 1, w.max_column, width)
        for split, df in frames.items():
            style_sheet(wb[f"Regimes_{split}"], len(df), len(keep), 20, columns=keep)

    print(f"excel  -> {xlsx.relative_to(config.PROJECT_ROOT)} "
          f"({xlsx.stat().st_size / 1e6:.1f} MB)")
    print(f"figures-> reports/figures/03_regime_timeline.png, 03_regime_bic.png")


if __name__ == "__main__":
    main()
