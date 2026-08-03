"""Step 2 — build the full feature matrix (RSI + MACD + ATR + ADX + Bollinger).

    python scripts/02_build_features.py

Reads   data/interim/nifty50_{train,test}.parquet     (step 1 output)
Writes  data/processed/features_{train,test}.parquet  (input to the HMM, step 3)
        reports/02_features.xlsx                      (the deliverable workbook)
        reports/figures/02_features_panel.png

`scripts/02a_rsi_features.py` still produces the RSI-only workbook; this script
supersedes it for pipeline purposes.
"""

from __future__ import annotations

import sys
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from src import config
from src.data.loader import load_clean
from src.features.pipeline import (
    FULL_WARMUP_BARS,
    INTERMEDIATE_COLS,
    RAW_SCALE_COLS,
    attach_all_features,
    model_feature_columns,
)

BODY_FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name=BODY_FONT, size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=BODY_FONT, size=13, bold=True, color="1F3864")
BOLD = Font(name=BODY_FONT, size=11, bold=True)
BODY = Font(name=BODY_FONT, size=10)
NOTE = Font(name=BODY_FONT, size=10, italic=True, color="555555")

# Columns excluded from the Excel export only (they bloat the file and add
# nothing a reader can use). The parquet keeps everything.
EXCEL_DROP = ["src_pivot", "src_200_EMA", "src_Supertrend(12,3)",
              "src_Supertrend(11,2)", "src_Supertrend(10,1)"]

BLOCK_NOTES = [
    ("RSI (7/14/21)",
     "Wilder RSI. Momentum: is the recent move overextended?",
     "Wilder RMA with an explicit simple-mean seed - a bare ewm() is not the "
     "same series and will not match TradingView."),
    ("MACD (12/26/9)",
     "Trend + momentum: fast EMA against slow EMA, and their signal line.",
     "Raw MACD is in index points and roughly doubles across the training "
     "window as Nifty triples. Feed macd_pct, never macd."),
    ("ATR (14)",
     "Volatility in true-range terms, including overnight gaps.",
     "Same scale problem as MACD - feed atr_pct. atr_pctile_1000 ranks it "
     "against the trailing 1000 bars so 2015 and 2024 are comparable."),
    ("ADX / DMI (14)",
     "Trend STRENGTH, unsigned. Separates trending from sideways.",
     "Already 0-100, no normalisation needed. ADX has no sign, so it must be "
     "paired with di_spread or a crash and a rally look identical."),
    ("Bollinger (20, 2sd)",
     "Position within the bands (%B) and how wide they are (bandwidth).",
     "Population std (ddof=0), per Bollinger's definition; pandas defaults to "
     "ddof=1, which widens every band by ~2.6%."),
]

ASSUMPTIONS = [
    "All indicators are computed on the CONTINUOUS series, not reset per "
    "session. Overnight gaps are genuine index moves, and a daily reset would "
    "discard 14 of the 25 bars in every session.",
    "The TEST split is warm-started from the last {warmup} bars of TRAIN, then "
    "those seed rows are discarded. This feeds past into future, which is the "
    "correct direction; the reverse would be leakage. attach_block() raises if "
    "the seed does not end strictly before the target begins.",
    "{warmup} bars is set by the two trailing-percentile columns "
    "(atr_pctile_1000, bb_bandwidth_pctile), which need a full ranking window "
    "of history - far more than the smoothing itself needs.",
    "is_warmup marks rows where ANY block is still undefined or seed-dominated. "
    "Drop those rows before training. Past that flag the matrix has zero NaNs "
    "and zero infinities, both asserted in tests.",
    "Columns in index points (macd, atr_14, bb_upper, ...) are kept for "
    "charting and chart-package validation but must NOT be fed to a model - "
    "they drift with the price level. Each has a scale-free counterpart; see "
    "the Columns sheet.",
    "Every block is cross-checked against the `ta` library and matches to "
    "machine precision beyond the warm-up region. RSI additionally reproduces "
    "Wilder's published worked example to within his table's rounding.",
]

FINDINGS = [
    "15.0% of closes fall outside the 2-sigma Bollinger bands, against ~5% "
    "expected under normality. Returns are fat-tailed - which is itself an "
    "argument for the multi-state HMM in step 3 over a single Gaussian.",
    "ADX > 25 flags 62% of bars as trending. Wilder calibrated that threshold "
    "on daily bars; on 15-minute data it is far less discriminating, so the "
    "threshold should be tuned against the HMM's own regime labels rather than "
    "taken as given.",
    "RSI(14) runs overbought on 17.9% of train bars vs oversold on 12.0%, "
    "reflecting a net strongly-upward 2015-2024 window. Expect RSI to lean "
    "bullish when the HMM labels regimes.",
    "Mean ATR% is 0.36 in March 2020 against 0.15 in June 2017, and Bollinger "
    "bandwidth is 5.79 vs 0.57 across the same two months - both blocks "
    "separate the COVID crash from a calm period by a wide margin.",
]


def summarise(df: pd.DataFrame, split: str, cols: list[str]) -> pd.DataFrame:
    usable = df.loc[~df["is_warmup"], cols]
    num = usable.select_dtypes("number")
    s = num.describe().T[["count", "mean", "std", "min", "50%", "max"]]
    s.insert(0, "split", split)
    s.insert(0, "feature", s.index)
    return s.reset_index(drop=True)


def write_readme(ws, meta: dict) -> None:
    for col, width in (("A", 24), ("B", 58), ("C", 62)):
        ws.column_dimensions[col].width = width

    r = 1
    ws.cell(r, 1, "Nifty 50 Capstone - Step 2: feature matrix").font = TITLE_FONT
    r += 1
    ws.cell(r, 1, "Generated by scripts/02_build_features.py. Values are computed "
                  "in Python, not by Excel formulas - this is a feature export, "
                  "not a model.").font = NOTE
    r += 2

    ws.cell(r, 1, "Coverage").font = BOLD
    r += 1
    for i, h in enumerate(["Sheet", "Rows", "Usable (post warm-up)", "Range"], start=1):
        c = ws.cell(r, i, h)
        c.font, c.fill = HEADER_FONT, HEADER_FILL
    r += 1
    for split, info in meta.items():
        ws.cell(r, 1, f"Features_{split}").font = BODY
        ws.cell(r, 2, info["rows"]).font = BODY
        ws.cell(r, 3, info["usable"]).font = BODY
        ws.cell(r, 4, info["range"]).font = BODY
        ws.cell(r, 2).number_format = "#,##0"
        ws.cell(r, 3).number_format = "#,##0"
        r += 1
    r += 1

    ws.cell(r, 1, "The five blocks").font = BOLD
    r += 1
    for i, h in enumerate(["Block", "What it measures", "Implementation note"], start=1):
        c = ws.cell(r, i, h)
        c.font, c.fill = HEADER_FONT, HEADER_FILL
    r += 1
    for name, what, note in BLOCK_NOTES:
        ws.cell(r, 1, name).font = Font(name=BODY_FONT, size=10, bold=True)
        for i, txt in ((2, what), (3, note)):
            c = ws.cell(r, i, txt)
            c.font, c.alignment = BODY, Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 42
        r += 1
    r += 1

    ws.cell(r, 1, "Method and assumptions").font = BOLD
    r += 1
    for line in ASSUMPTIONS:
        ws.cell(r, 1, "-").font = BODY
        c = ws.cell(r, 2, line.format(warmup=FULL_WARMUP_BARS))
        c.font, c.alignment = BODY, Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws.row_dimensions[r].height = 46
        r += 1
    r += 1

    ws.cell(r, 1, "Findings worth reporting").font = BOLD
    r += 1
    for line in FINDINGS:
        ws.cell(r, 1, "-").font = BODY
        c = ws.cell(r, 2, line)
        c.font, c.alignment = BODY, Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws.row_dimensions[r].height = 46
        r += 1
    r += 1

    ws.cell(r, 1, "Known limitation").font = BOLD
    r += 1
    c = ws.cell(r, 2, "The source data has no volume column, so OBV, VWAP "
                      "deviation and volume-confirmed breakouts are unavailable. "
                      "All five blocks above need only OHLC and are unaffected.")
    c.font, c.alignment = NOTE, Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws.row_dimensions[r].height = 46


def write_columns_sheet(ws, df: pd.DataFrame, model_cols: set[str]) -> None:
    for col, width in (("A", 26), ("B", 14), ("C", 16), ("D", 46)):
        ws.column_dimensions[col].width = width
    for i, h in enumerate(["Column", "Block", "Model input?", "Note"], start=1):
        c = ws.cell(1, i, h)
        c.font, c.fill = HEADER_FONT, HEADER_FILL
    ws.freeze_panes = "A2"

    def block_of(name: str) -> str:
        for prefix, label in (("rsi", "RSI"), ("macd", "MACD"),
                              ("atr", "ATR"), ("true_range", "ATR"),
                              ("tr_over", "ATR"), ("adx", "ADX"),
                              ("di_", "ADX"), ("dx", "ADX"), ("bb_", "Bollinger")):
            if name.startswith(prefix):
                return label
        return "step 1"

    r = 2
    for name in df.columns:
        is_model = name in model_cols
        note = ""
        if name in RAW_SCALE_COLS:
            note = f"In index points - drifts with price level. Use {RAW_SCALE_COLS[name]}."
        elif name in INTERMEDIATE_COLS:
            note = f"Intermediate step, redundant with {INTERMEDIATE_COLS[name]}."
        elif name.endswith("_is_warmup") or name == "is_warmup":
            note = "Bookkeeping: drop rows where this is TRUE before training."
        elif name.startswith("src_"):
            note = "Vendor column, provenance unknown. Reference only."
        ws.cell(r, 1, name).font = BODY
        ws.cell(r, 2, block_of(name)).font = BODY
        ws.cell(r, 3, "yes" if is_model else "no").font = BODY
        c = ws.cell(r, 4, note)
        c.font, c.alignment = BODY, Alignment(wrap_text=True, vertical="top")
        r += 1


def style_data_sheet(ws, n_rows: int, cols: list[str]) -> None:
    for j, name in enumerate(cols, start=1):
        letter = get_column_letter(j)
        ws.column_dimensions[letter].width = max(11, min(len(name) + 3, 20))
        c = ws.cell(1, j)
        c.font, c.fill = HEADER_FONT, HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if name == "timestamp":
            for i in range(2, n_rows + 2):
                ws.cell(i, j).number_format = "yyyy-mm-dd hh:mm"
        elif name in {"open", "high", "low", "close"}:
            for i in range(2, n_rows + 2):
                ws.cell(i, j).number_format = "#,##0.00"
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{n_rows + 1}"
    ws.row_dimensions[1].height = 32


def plot_panel(df: pd.DataFrame, out: Path, n: int = 600) -> None:
    """One panel per block, on a positional axis so overnight gaps don't stretch."""
    d = df.tail(n).reset_index(drop=True)
    x = np.arange(len(d))

    fig, ax = plt.subplots(6, 1, figsize=(13, 15), sharex=True,
                           gridspec_kw={"height_ratios": [2, 1, 1, 1, 1, 1]})

    ax[0].plot(x, d["close"], lw=0.9, color="#1f3864")
    ax[0].plot(x, d["bb_upper"], lw=0.6, color="#a0aec0")
    ax[0].plot(x, d["bb_lower"], lw=0.6, color="#a0aec0")
    ax[0].fill_between(x, d["bb_lower"], d["bb_upper"], color="#a0aec0", alpha=0.18)
    ax[0].set_ylabel("Close + BB(20,2)")
    ax[0].set_title(f"Step 2 feature matrix — last {n:,} train bars "
                    f"({d['session_date'].iloc[0]} to {d['session_date'].iloc[-1]})")

    ax[1].plot(x, d["rsi_14"], lw=1.0, color="#c05621")
    for lvl, col in ((70, "#c53030"), (30, "#2b6cb0")):
        ax[1].axhline(lvl, color=col, ls="--", lw=0.8)
    ax[1].set_ylim(0, 100)
    ax[1].set_ylabel("RSI(14)")

    ax[2].plot(x, d["macd_pct"], lw=0.9, color="#2b6cb0", label="MACD %")
    ax[2].plot(x, d["macd_signal_pct"], lw=0.8, color="#c05621", label="signal")
    ax[2].bar(x, d["macd_hist_pct"], color="#a0aec0", width=1.0)
    ax[2].axhline(0, color="#4a5568", lw=0.6)
    ax[2].set_ylabel("MACD (% of close)")
    ax[2].legend(loc="upper left", fontsize=8, ncol=2)

    ax[3].plot(x, d["atr_pct"], lw=0.9, color="#2f855a")
    ax[3].set_ylabel("ATR (% of close)")

    ax[4].plot(x, d["adx_14"], lw=1.0, color="#6b46c1", label="ADX")
    ax[4].plot(x, d["di_plus"], lw=0.6, color="#2f855a", label="+DI")
    ax[4].plot(x, d["di_minus"], lw=0.6, color="#c53030", label="-DI")
    ax[4].axhline(25, color="#a0aec0", ls="--", lw=0.8)
    ax[4].set_ylabel("ADX / DMI")
    ax[4].legend(loc="upper left", fontsize=8, ncol=3)

    ax[5].plot(x, d["bb_pct_b"], lw=0.9, color="#b7791f")
    ax[5].axhline(1, color="#c53030", ls="--", lw=0.8)
    ax[5].axhline(0, color="#2b6cb0", ls="--", lw=0.8)
    ax[5].set_ylabel("Bollinger %B")

    opens = d.index[d["is_session_open"]].tolist()
    step = max(1, len(opens) // 12)
    ticks = opens[::step]
    for a in ax:
        a.grid(alpha=0.25)
        a.set_xlim(0, len(d) - 1)
        for t in opens:
            a.axvline(t, color="#edf2f7", lw=0.4, zorder=0)
    ax[-1].set_xticks(ticks)
    ax[-1].set_xticklabels([d["session_date"].iloc[t] for t in ticks],
                           rotation=45, ha="right", fontsize=8)

    fig.tight_layout()
    out.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out, dpi=120)
    plt.close(fig)


def main() -> None:
    config.PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    config.REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    train = load_clean("train")
    test = load_clean("test")

    frames = {
        "train": attach_all_features(train),
        "test": attach_all_features(test, seed=train),
    }

    model_cols = model_feature_columns(frames["train"])
    summaries, meta = [], {}

    for split, df in frames.items():
        out = config.PROCESSED_DIR / f"features_{split}.parquet"
        df.to_parquet(out, index=False)
        usable = int((~df["is_warmup"]).sum())
        meta[split] = {
            "rows": len(df),
            "usable": usable,
            "range": f"{df['timestamp'].iloc[0]:%Y-%m-%d} .. {df['timestamp'].iloc[-1]:%Y-%m-%d}",
        }
        summaries.append(summarise(df, split, model_cols))
        print(f"{split:5s}  {len(df):>7,} bars  {df.shape[1]:>3} cols  "
              f"usable {usable:>7,}  -> {out.name}")

    print(f"model-input columns: {len(model_cols)}")

    xlsx = config.REPORTS_DIR / "02_features.xlsx"
    print(f"writing {xlsx.name} ...")
    excel_cols = [c for c in frames["train"].columns if c not in EXCEL_DROP]

    with pd.ExcelWriter(xlsx, engine="openpyxl") as xw:
        pd.DataFrame({"_": []}).to_excel(xw, sheet_name="README", index=False)
        pd.DataFrame({"_": []}).to_excel(xw, sheet_name="Columns", index=False)
        pd.concat(summaries, ignore_index=True).to_excel(
            xw, sheet_name="Summary", index=False)
        for split, df in frames.items():
            ex = df[excel_cols].copy()
            ex["timestamp"] = ex["timestamp"].dt.tz_localize(None)  # Excel has no tz
            ex.to_excel(xw, sheet_name=f"Features_{split}", index=False)

        wb = xw.book
        wb._named_styles["Normal"].font = Font(name=BODY_FONT, size=10)

        ws = wb["README"]
        ws.delete_rows(1, ws.max_row)
        write_readme(ws, meta)

        ws = wb["Columns"]
        ws.delete_rows(1, ws.max_row)
        write_columns_sheet(ws, frames["train"][excel_cols], set(model_cols))

        ws = wb["Summary"]
        for j in range(1, ws.max_column + 1):
            c = ws.cell(1, j)
            c.font, c.fill = HEADER_FONT, HEADER_FILL
            ws.column_dimensions[get_column_letter(j)].width = 26 if j == 1 else 14
        for i in range(2, ws.max_row + 1):
            for j in range(4, ws.max_column + 1):
                ws.cell(i, j).number_format = "0.0000"
        ws.freeze_panes = "A2"

        for split, df in frames.items():
            style_data_sheet(wb[f"Features_{split}"], len(df), excel_cols)

    fig_path = config.REPORTS_DIR / "figures" / "02_features_panel.png"
    plot_panel(frames["train"], fig_path)

    size_mb = xlsx.stat().st_size / 1e6
    print(f"excel  -> {xlsx.relative_to(config.PROJECT_ROOT)}  ({size_mb:.1f} MB)")
    print(f"figure -> {fig_path.relative_to(config.PROJECT_ROOT)}")


if __name__ == "__main__":
    main()
