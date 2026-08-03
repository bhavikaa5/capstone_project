"""Step 2, part 1 — build the RSI feature block and export it to Excel.

    python scripts/02a_rsi_features.py

Reads   data/interim/nifty50_{train,test}.parquet   (step 1 output)
Writes  data/processed/rsi_{train,test}.parquet     (for the rest of the pipeline)
        reports/02_rsi_features.xlsx                (the deliverable workbook)
        reports/figures/02_rsi_train.png
"""

from __future__ import annotations

import sys
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from src import config
from src.data.loader import load_clean
from src.features import rsi as rsi_mod

BODY_FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name=BODY_FONT, size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=BODY_FONT, size=13, bold=True, color="1F3864")
NOTE_FONT = Font(name=BODY_FONT, size=10, italic=True, color="555555")

# column -> (excel number format, width)
FORMATS = {
    "timestamp": ("yyyy-mm-dd hh:mm", 18),
    "open": ("#,##0.00", 11),
    "high": ("#,##0.00", 11),
    "low": ("#,##0.00", 11),
    "close": ("#,##0.00", 11),
    "rsi_7": ("0.00", 9),
    "rsi_14": ("0.00", 9),
    "rsi_21": ("0.00", 9),
    "rsi_14_norm": ("0.0000", 12),
    "rsi_14_chg": ("0.0000", 11),
    "rsi_14_slope_5": ("0.0000", 13),
    "rsi_spread_7_21": ("0.0000", 14),
    "rsi_14_zone": ("0", 11),
}
EXPORT_COLS = [
    "timestamp", "session_date", "bar_of_day",
    "open", "high", "low", "close",
    "rsi_7", "rsi_14", "rsi_21",
    "rsi_14_norm", "rsi_14_chg", "rsi_14_slope_5", "rsi_spread_7_21",
    "rsi_14_overbought", "rsi_14_oversold", "rsi_14_zone", "rsi_is_warmup",
]

DEFINITIONS = [
    ("timestamp", "Bar open time, Asia/Kolkata (tz dropped for Excel)", "step 1"),
    ("session_date", "Trading date the bar belongs to", "step 1"),
    ("bar_of_day", "0-24 index within the session", "step 1"),
    ("open / high / low / close", "Nifty 50 index price for the 15-min bar", "raw data"),
    ("rsi_7", "Wilder RSI, period 7 (fast momentum)", "100 - 100/(1+RS)"),
    ("rsi_14", "Wilder RSI, period 14 - the canonical setting", "100 - 100/(1+RS)"),
    ("rsi_21", "Wilder RSI, period 21 (slow momentum)", "100 - 100/(1+RS)"),
    ("rsi_14_norm", "rsi_14 centred and scaled to [-1, 1]", "(rsi_14 - 50) / 50"),
    ("rsi_14_chg", "One-bar change in rsi_14 (momentum direction)", "rsi_14 - rsi_14[t-1]"),
    ("rsi_14_slope_5", "5-bar slope of rsi_14 (smoothed direction)", "(rsi_14 - rsi_14[t-5]) / 5"),
    ("rsi_spread_7_21", "Fast minus slow RSI - divergence proxy", "rsi_7 - rsi_21"),
    ("rsi_14_overbought", "TRUE when rsi_14 > 70", "Wilder threshold"),
    ("rsi_14_oversold", "TRUE when rsi_14 < 30", "Wilder threshold"),
    ("rsi_14_zone", "+1 overbought, 0 neutral, -1 oversold", "overbought - oversold"),
    ("rsi_is_warmup", "TRUE while the smoothing is still seed-dominated", "see Method"),
]

METHOD = [
    "RS = Wilder RMA(gains, n) / Wilder RMA(losses, n), where Wilder's RMA uses "
    "alpha = 1/n and is seeded with the simple mean of the first n price changes.",
    "A plain pandas ewm() is NOT equivalent - it adapts faster and will not match "
    "TradingView or the vendor's own indicator columns.",
    "Computed on the continuous close series, not reset per session. Overnight gaps "
    "are genuine index moves, and resetting each morning would discard 14 of the 25 "
    "bars in every session.",
    "Special cases: an unbroken run of up-bars gives avg_loss = 0 -> RSI = 100; a "
    "perfectly flat window gives 0/0 -> RSI = 50.",
    "The first 14 bars of the TRAIN split are NaN by definition - RSI is undefined "
    "until 14 price changes exist. rsi_is_warmup also flags the first 105 bars "
    "(5 x the longest period), where the seed still measurably influences the value.",
    "The TEST split is warm-started from the last 250 bars of TRAIN, so it has no "
    "NaNs and no warm-up bias. This feeds past into future, which is the correct "
    "direction; the reverse would be leakage.",
    "Validation: matches the `ta` library to machine precision beyond bar 1000 "
    "(all disagreement is confined to the warm-up region, where `ta` seeds off a "
    "13-change window), and reproduces Wilder's published worked example to within "
    "0.07, which is the rounding in his printed table.",
]


def summarise(df: pd.DataFrame, split: str) -> pd.DataFrame:
    """Per-split descriptive stats for the RSI columns."""
    cols = ["rsi_7", "rsi_14", "rsi_21", "rsi_14_chg", "rsi_spread_7_21"]
    s = df[cols].describe().T[["count", "mean", "std", "min", "50%", "max"]]
    s.insert(0, "split", split)
    s.insert(0, "feature", s.index)
    ob = df["rsi_14_overbought"].mean() * 100
    os_ = df["rsi_14_oversold"].mean() * 100
    extra = pd.DataFrame([
        {"feature": "% bars overbought (rsi_14>70)", "split": split, "mean": ob},
        {"feature": "% bars oversold (rsi_14<30)", "split": split, "mean": os_},
        {"feature": "% bars flagged warm-up", "split": split,
         "mean": df["rsi_is_warmup"].mean() * 100},
    ])
    return pd.concat([s.reset_index(drop=True), extra], ignore_index=True)


def style_data_sheet(ws, n_rows: int, cols: list[str]) -> None:
    thin = Side(style="thin", color="D0D0D0")
    for j, name in enumerate(cols, start=1):
        letter = get_column_letter(j)
        fmt, width = FORMATS.get(name, (None, max(12, len(name) + 2)))
        ws.column_dimensions[letter].width = width
        cell = ws.cell(row=1, column=j)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
        if fmt:
            for i in range(2, n_rows + 2):
                ws.cell(row=i, column=j).number_format = fmt
    ws.freeze_panes = "E2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{n_rows + 1}"
    ws.row_dimensions[1].height = 30


def write_readme(ws, reports: dict[str, dict]) -> None:
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 62
    ws.column_dimensions["C"].width = 46

    r = 1
    ws.cell(r, 1, "Nifty 50 Capstone - Step 2 (part 1): RSI feature block").font = TITLE_FONT
    r += 1
    ws.cell(r, 1, "Generated by scripts/02a_rsi_features.py from the step-1 cleaned "
                  "data. Values are computed in Python, not by Excel formulas - this "
                  "sheet is a feature export, not a model.").font = NOTE_FONT
    r += 2

    ws.cell(r, 1, "Coverage").font = Font(name=BODY_FONT, size=11, bold=True)
    r += 1
    for h, c in zip(["Sheet", "Rows", "Range"], range(1, 4)):
        cell = ws.cell(r, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    r += 1
    for split, info in reports.items():
        ws.cell(r, 1, f"RSI_{split}").font = Font(name=BODY_FONT, size=10)
        ws.cell(r, 2, info["rows"]).font = Font(name=BODY_FONT, size=10)
        ws.cell(r, 2).number_format = "#,##0"
        ws.cell(r, 3, info["range"]).font = Font(name=BODY_FONT, size=10)
        r += 1
    r += 1

    ws.cell(r, 1, "Column definitions").font = Font(name=BODY_FONT, size=11, bold=True)
    r += 1
    for h, c in zip(["Column", "Meaning", "Formula / source"], range(1, 4)):
        cell = ws.cell(r, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    r += 1
    for name, meaning, formula in DEFINITIONS:
        ws.cell(r, 1, name).font = Font(name=BODY_FONT, size=10, bold=True)
        ws.cell(r, 2, meaning).font = Font(name=BODY_FONT, size=10)
        ws.cell(r, 3, formula).font = Font(name=BODY_FONT, size=10)
        ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(r, 3).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    r += 1

    ws.cell(r, 1, "Method, assumptions and validation").font = Font(
        name=BODY_FONT, size=11, bold=True)
    r += 1
    for line in METHOD:
        ws.cell(r, 1, "-").font = Font(name=BODY_FONT, size=10)
        c = ws.cell(r, 2, line)
        c.font = Font(name=BODY_FONT, size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 42
        r += 1
    r += 1

    ws.cell(r, 1, "Known limitation").font = Font(name=BODY_FONT, size=11, bold=True)
    r += 1
    c = ws.cell(r, 2, "The source data has no volume column, so no volume-confirmed "
                      "RSI variant (e.g. RSI on OBV) is available. RSI itself needs "
                      "only close prices and is unaffected.")
    c.font = NOTE_FONT
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 42


def plot_rsi(df: pd.DataFrame, out: Path, n: int = 600) -> None:
    """Plot against bar position, not wall-clock time.

    A datetime axis renders every overnight and weekend gap as a long diagonal,
    which on 15-minute data drowns out the signal. Ticks are relabelled with the
    session date so the axis still reads as calendar time.
    """
    d = df.tail(n).reset_index(drop=True)
    x = np.arange(len(d))

    fig, ax = plt.subplots(2, 1, figsize=(13, 7), sharex=True,
                           gridspec_kw={"height_ratios": [2, 1]})
    ax[0].plot(x, d["close"], lw=0.9, color="#1f3864")
    ax[0].set_ylabel("Nifty 50 close")
    ax[0].set_title(f"RSI feature block — last {n:,} train bars "
                    f"({d['session_date'].iloc[0]} to {d['session_date'].iloc[-1]})")
    ax[0].grid(alpha=0.25)

    ax[1].plot(x, d["rsi_7"], lw=0.6, color="#a0aec0", label="RSI(7)")
    ax[1].plot(x, d["rsi_14"], lw=1.1, color="#c05621", label="RSI(14)")
    ax[1].plot(x, d["rsi_21"], lw=0.8, color="#2f855a", label="RSI(21)")
    ax[1].axhline(70, color="#c53030", ls="--", lw=0.8)
    ax[1].axhline(30, color="#2b6cb0", ls="--", lw=0.8)
    ax[1].axhline(50, color="#cbd5e0", lw=0.6)
    ax[1].fill_between(x, 70, 100, where=d["rsi_14"] > 70, color="#c53030", alpha=0.15)
    ax[1].fill_between(x, 0, 30, where=d["rsi_14"] < 30, color="#2b6cb0", alpha=0.15)
    ax[1].set_ylim(0, 100)
    ax[1].set_ylabel("RSI")
    ax[1].set_xlim(0, len(d) - 1)
    ax[1].legend(loc="upper left", ncol=3, fontsize=8)
    ax[1].grid(alpha=0.25)

    # One tick per session start, thinned to ~12 labels.
    opens = d.index[d["is_session_open"]].tolist()
    step = max(1, len(opens) // 12)
    ticks = opens[::step]
    ax[1].set_xticks(ticks)
    ax[1].set_xticklabels([d["session_date"].iloc[t] for t in ticks],
                          rotation=45, ha="right", fontsize=8)
    for t in opens:                      # faint session separators
        for a in ax:
            a.axvline(t, color="#e2e8f0", lw=0.4, zorder=0)

    fig.tight_layout()
    out.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out, dpi=130)
    plt.close(fig)


def main() -> None:
    config.PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    config.REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    train = load_clean("train")
    test = load_clean("test")

    train_f = rsi_mod.attach_rsi(train)
    test_f = rsi_mod.attach_rsi(test, seed=train)   # warm-start from train tail

    frames = {"train": train_f, "test": test_f}
    summaries, meta = [], {}

    for split, df in frames.items():
        out = config.PROCESSED_DIR / f"rsi_{split}.parquet"
        df.to_parquet(out, index=False)
        summaries.append(summarise(df, split))
        meta[split] = {
            "rows": len(df),
            "range": f"{df['timestamp'].iloc[0]:%Y-%m-%d} .. {df['timestamp'].iloc[-1]:%Y-%m-%d}",
        }
        nan = int(df["rsi_14"].isna().sum())
        print(f"{split:5s}  {len(df):>7,} bars   rsi_14 NaN: {nan:>3}   "
              f"warm-up flagged: {int(df['rsi_is_warmup'].sum()):>3}   -> {out.name}")

    xlsx = config.REPORTS_DIR / "02_rsi_features.xlsx"
    print(f"writing {xlsx.name} (this takes ~30s for {len(train_f):,} rows)...")

    with pd.ExcelWriter(xlsx, engine="openpyxl") as xw:
        pd.DataFrame({"placeholder": []}).to_excel(xw, sheet_name="README", index=False)
        pd.concat(summaries, ignore_index=True).to_excel(
            xw, sheet_name="Summary", index=False)
        for split, df in frames.items():
            ex = df[EXPORT_COLS].copy()
            ex["timestamp"] = ex["timestamp"].dt.tz_localize(None)  # Excel has no tz
            ex.to_excel(xw, sheet_name=f"RSI_{split}", index=False)

        wb = xw.book
        wb._named_styles["Normal"].font = Font(name=BODY_FONT, size=10)

        ws = wb["README"]
        ws.delete_rows(1, ws.max_row)
        write_readme(ws, meta)

        ws = wb["Summary"]
        for j in range(1, ws.max_column + 1):
            cell = ws.cell(1, j)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            ws.column_dimensions[get_column_letter(j)].width = 30 if j == 1 else 14
        for i in range(2, ws.max_row + 1):
            for j in range(3, ws.max_column + 1):
                ws.cell(i, j).number_format = "0.0000"
        ws.freeze_panes = "A2"

        for split, df in frames.items():
            style_data_sheet(wb[f"RSI_{split}"], len(df), EXPORT_COLS)

    fig_path = config.REPORTS_DIR / "figures" / "02_rsi_train.png"
    plot_rsi(train_f, fig_path)

    print(f"excel  -> {xlsx.relative_to(config.PROJECT_ROOT)}")
    print(f"figure -> {fig_path.relative_to(config.PROJECT_ROOT)}")


if __name__ == "__main__":
    main()
