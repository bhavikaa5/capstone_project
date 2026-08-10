"""Step 4 — train and compare the three architectures.

    python scripts/04_train_models.py [--epochs 40] [--quick]

Reads   data/processed/regimes_{train,test}.parquet   (step 3 output)
Writes  models/{name}.keras                           (trained weights)
        models/scaler.npz, models/feature_columns.json
        data/processed/predictions_test.parquet
        reports/04_models.xlsx
        reports/figures/04_training_curves.png, 04_model_comparison.png
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import time
from pathlib import Path

os.environ.setdefault("TF_CPP_MIN_LOG_LEVEL", "3")

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import keras
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src import config
from src.models.architectures import ARCHITECTURES, set_seeds
from src.models.dataset import (
    HORIZON,
    LOOKBACK,
    Scaler,
    baseline_scores,
    build_sequences,
    chronological_split,
    feature_columns,
)
from src.models.evaluate import (
    classification_metrics,
    cost_sensitivity,
    directional_strategy,
    persistence_agreement,
    pick_threshold,
    regime_breakdown,
)

BODY_FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name=BODY_FONT, size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=BODY_FONT, size=13, bold=True, color="1F3864")
BOLD = Font(name=BODY_FONT, size=11, bold=True)
BODY = Font(name=BODY_FONT, size=10)
NOTE = Font(name=BODY_FONT, size=10, italic=True, color="555555")

COLOURS = {"lstm": "#2b6cb0", "cnn_bilstm_attention": "#c05621",
           "transformer": "#2f855a"}

METHOD = [
    "TARGET: direction of the next 15-minute bar. 1 if log close rises over the "
    "next bar, else 0. The window ending at bar t contains bars t-31..t only; "
    "nothing after t enters it.",
    "SPLIT: the last 15% of the training file becomes validation, "
    "chronologically, with a purge gap of 33 windows between the two. Without "
    "the gap the final training windows share input bars with the first "
    "validation windows and a training label reaches into the validation "
    "period. The test file is never touched during training.",
    "SCALING: per-feature standardisation fit on the TRAINING windows only, "
    "then applied frozen to validation and test.",
    "THRESHOLD: chosen on validation by accuracy, then frozen and applied to "
    "test. Tuning it on test would be scoring against the answer key.",
    "CAUSALITY IN THE ARCHITECTURES: the CNN uses padding='causal', not 'same' "
    "- 'same' padding lets the convolution at bar t read bars t+1 and t+2. The "
    "transformer applies a causal attention mask for the same reason.",
    "Early stopping monitors validation AUC with patience 6 and restores the "
    "best weights, so the reported model is the best-validating epoch rather "
    "than the last one.",
]

FINDINGS_INTRO = (
    "ACCURACY ALONE IS NOT A RESULT HERE. The classes are nearly balanced "
    "(50.8% up on test), so a coin flip scores ~50%. More importantly, "
    "15-minute Nifty returns have lag-1 autocorrelation of 0.18, so the "
    "one-line rule 'the next bar goes the same way as this one' already scores "
    "61.4% on test. Any model below that number is worse than one line of "
    "code. Every table below reports lift over persistence for that reason."
)


def build_data(quick: bool = False):
    train_df = pd.read_parquet(config.PROCESSED_DIR / "regimes_train.parquet")
    test_df = pd.read_parquet(config.PROCESSED_DIR / "regimes_test.parquet")

    cols = feature_columns(train_df)
    full = build_sequences(train_df, cols, LOOKBACK, HORIZON)
    train, val = chronological_split(full)
    test = build_sequences(test_df, cols, LOOKBACK, HORIZON)

    if quick:                      # smoke path: a fifth of the data, few epochs
        train.X, train.y = train.X[-9000:], train.y[-9000:]
        train.row_index = train.row_index[-9000:]

    scaler = Scaler().fit(train.X)
    for part in (train, val, test):
        part.X = scaler.transform(part.X)

    baselines = {
        "train": baseline_scores(train_df, train),
        "val": baseline_scores(train_df, val),
        "test": baseline_scores(test_df, test),
    }
    return train_df, test_df, cols, train, val, test, scaler, baselines


def train_one(name, builder, train, val, epochs, batch_size):
    set_seeds(42)
    model = builder((train.X.shape[1], train.X.shape[2]))
    callbacks = [
        keras.callbacks.EarlyStopping(monitor="val_auc", mode="max", patience=6,
                                      restore_best_weights=True, verbose=0),
        keras.callbacks.ReduceLROnPlateau(monitor="val_auc", mode="max", factor=0.5,
                                          patience=3, min_lr=1e-5, verbose=0),
    ]
    t0 = time.time()
    history = model.fit(
        train.X, train.y,
        validation_data=(val.X, val.y),
        epochs=epochs, batch_size=batch_size,
        callbacks=callbacks, verbose=0, shuffle=True,
    )
    elapsed = time.time() - t0
    best_epoch = int(np.argmax(history.history["val_auc"])) + 1
    print(f"  {name:24s} {len(history.history['loss']):>3} epochs "
          f"(best {best_epoch:>2})  {elapsed / 60:>5.1f} min  "
          f"params={model.count_params():,}")
    return model, history.history, elapsed


def plot_curves(histories: dict, out: Path) -> None:
    fig, ax = plt.subplots(1, 3, figsize=(15, 4.2))
    for name, h in histories.items():
        c = COLOURS[name]
        e = range(1, len(h["loss"]) + 1)
        ax[0].plot(e, h["loss"], color=c, lw=1.2, label=f"{name} train")
        ax[0].plot(e, h["val_loss"], color=c, lw=1.2, ls="--")
        ax[1].plot(e, h["val_acc"], color=c, lw=1.4, label=name)
        ax[2].plot(e, h["val_auc"], color=c, lw=1.4, label=name)
    ax[0].set_title("loss (solid = train, dashed = val)")
    ax[1].set_title("validation accuracy")
    ax[2].set_title("validation AUC")
    ax[2].axhline(0.5, color="#a0aec0", ls=":", lw=1)
    for a, lab in zip(ax, ["binary cross-entropy", "accuracy", "AUC"]):
        a.set_xlabel("epoch")
        a.set_ylabel(lab)
        a.grid(alpha=0.25)
        a.legend(fontsize=8)
    fig.tight_layout()
    out.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out, dpi=130)
    plt.close(fig)


def plot_comparison(summary: pd.DataFrame, baselines: dict, out: Path) -> None:
    test = summary[summary["split"] == "test"].set_index("model")
    names = list(test.index)
    x = np.arange(len(names))

    fig, ax = plt.subplots(1, 2, figsize=(13, 4.8))

    ax[0].bar(x, test["accuracy"] * 100, color=[COLOURS[n] for n in names], width=0.55)
    ax[0].axhline(baselines["test"]["persistence"] * 100, color="#c53030", ls="--", lw=1.6,
                  label=f"persistence {baselines['test']['persistence']*100:.1f}%")
    ax[0].axhline(baselines["test"]["majority"] * 100, color="#718096", ls=":", lw=1.4,
                  label=f"majority {baselines['test']['majority']*100:.1f}%")
    ax[0].set_xticks(x)
    ax[0].set_xticklabels([n.replace("_", "\n") for n in names], fontsize=8)
    ax[0].set_ylabel("test accuracy (%)")
    ax[0].set_ylim(45, max(70, test["accuracy"].max() * 100 + 4))
    ax[0].set_title("Test accuracy vs the baselines that matter")
    ax[0].legend(fontsize=8)
    for xi, v in zip(x, test["accuracy"] * 100):
        ax[0].text(xi, v + 0.4, f"{v:.2f}", ha="center", fontsize=9)

    ax[1].bar(x, test["mcc"], color=[COLOURS[n] for n in names], width=0.55)
    ax[1].axhline(0, color="#4a5568", lw=1)
    ax[1].set_xticks(x)
    ax[1].set_xticklabels([n.replace("_", "\n") for n in names], fontsize=8)
    ax[1].set_ylabel("Matthews correlation coefficient")
    ax[1].set_title("MCC — collapses to 0 for a constant predictor")
    for xi, v in zip(x, test["mcc"]):
        ax[1].text(xi, v + (0.004 if v >= 0 else -0.012), f"{v:.3f}",
                   ha="center", fontsize=9)

    for a in ax:
        a.grid(alpha=0.25, axis="y")
    fig.tight_layout()
    out.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out, dpi=130)
    plt.close(fig)


def write_readme(ws, meta: dict, baselines: dict, summary: pd.DataFrame) -> None:
    for col, width in (("A", 24), ("B", 58), ("C", 56)):
        ws.column_dimensions[col].width = width
    r = 1
    ws.cell(r, 1, "Nifty 50 Capstone - Step 4: model training and accuracy").font = TITLE_FONT
    r += 2

    ws.cell(r, 1, "READ THIS FIRST").font = BOLD
    r += 1
    c = ws.cell(r, 2, FINDINGS_INTRO)
    c.font, c.alignment = Font(name=BODY_FONT, size=10, bold=True, color="C53030"), \
        Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws.row_dimensions[r].height = 96
    r += 2

    ws.cell(r, 1, "Setup").font = BOLD
    r += 1
    for label, value in [
        ("Target", f"direction of next {HORIZON} bar(s)"),
        ("Lookback window", f"{LOOKBACK} bars"),
        ("Features", meta["n_features"]),
        ("Train windows", f"{meta['n_train']:,}"),
        ("Validation windows", f"{meta['n_val']:,}"),
        ("Test windows", f"{meta['n_test']:,}"),
        ("Purge gap", f"{LOOKBACK + HORIZON} windows"),
    ]:
        ws.cell(r, 1, label).font = Font(name=BODY_FONT, size=10, bold=True)
        ws.cell(r, 2, str(value)).font = BODY
        r += 1
    r += 1

    ws.cell(r, 1, "Baselines to beat").font = BOLD
    r += 1
    for i, h in enumerate(["Split", "Majority class", "Persistence"], start=1):
        c = ws.cell(r, i, h)
        c.font, c.fill = HEADER_FONT, HEADER_FILL
    r += 1
    for split, b in baselines.items():
        ws.cell(r, 1, split).font = BODY
        ws.cell(r, 2, b["majority"]).font = BODY
        ws.cell(r, 3, b["persistence"]).font = BODY
        for j in (2, 3):
            ws.cell(r, j).number_format = "0.00%"
        r += 1
    r += 1

    ws.cell(r, 1, "Result").font = BOLD
    r += 1
    test = summary[summary["split"] == "test"]
    best = test.loc[test["accuracy"].idxmax()]
    beat = test["accuracy"].max() > baselines["test"]["persistence"]
    verdict = (
        f"Best test accuracy: {best['model']} at {best['accuracy']*100:.2f}%, "
        f"against a persistence baseline of {baselines['test']['persistence']*100:.2f}%. "
        + ("The model beats the baseline."
           if beat else
           "NO MODEL BEATS THE PERSISTENCE BASELINE. Reported as-is: on this "
           "target and horizon the architectures do not add value over a "
           "one-line momentum rule. That is a finding, not a bug, and it is the "
           "honest thing to put in the report.")
    )
    c = ws.cell(r, 2, verdict)
    c.font = Font(name=BODY_FONT, size=10, bold=True,
                  color="2F855A" if beat else "C53030")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws.row_dimensions[r].height = 80
    r += 2

    ws.cell(r, 1, "Method and leakage controls").font = BOLD
    r += 1
    for line in METHOD:
        ws.cell(r, 1, "-").font = BODY
        c = ws.cell(r, 2, line)
        c.font, c.alignment = BODY, Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws.row_dimensions[r].height = 58
        r += 1


def style(ws, n_rows: int, n_cols: int, first_width: int = 24) -> None:
    for j in range(1, n_cols + 1):
        c = ws.cell(1, j)
        c.font, c.fill = HEADER_FONT, HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = first_width if j == 1 else 14
    ws.freeze_panes = "B2"
    ws.row_dimensions[1].height = 30


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--epochs", type=int, default=40)
    ap.add_argument("--batch-size", type=int, default=512)
    ap.add_argument("--quick", action="store_true", help="smoke run on a subset")
    ap.add_argument("--report-only", action="store_true",
                    help="reload saved models and rebuild the report without retraining")
    args = ap.parse_args()
    if args.quick:
        args.epochs = min(args.epochs, 3)

    models_dir = config.PROJECT_ROOT / "models"
    models_dir.mkdir(exist_ok=True)

    train_df, test_df, cols, train, val, test, scaler, baselines = build_data(args.quick)
    print(f"features={len(cols)}  train={len(train):,}  val={len(val):,}  test={len(test):,}")
    print("baselines (accuracy to beat):")
    for split, b in baselines.items():
        print(f"  {split:5s} majority={b['majority']*100:.2f}%  "
              f"persistence={b['persistence']*100:.2f}%")

    print(f"\ntraining {len(ARCHITECTURES)} architectures "
          f"(max {args.epochs} epochs, batch {args.batch_size}):")

    trained, histories, rows, regime_rows, strat_rows = {}, {}, [], [], []
    cost_rows, agree_rows = [], []

    log_close = np.log(test_df["close"].to_numpy())
    fwd_all = np.full(len(test_df), np.nan)
    fwd_all[:-HORIZON] = log_close[HORIZON:] - log_close[:-HORIZON]
    fwd_test = fwd_all[test.row_index]
    prev_ret_test = test_df.reset_index(drop=True)["ret_1"].to_numpy()[test.row_index]

    for name, builder in ARCHITECTURES.items():
        if args.report_only:
            model = keras.models.load_model(models_dir / f"{name}.keras")
            # Runs made before histories were persisted have no curve data; the
            # report is still fully reproducible without it, so carry on and
            # leave the existing training-curves figure in place.
            hist_path = models_dir / f"{name}_history.json"
            hist = json.loads(hist_path.read_text()) if hist_path.exists() else None
            elapsed = 0.0
            print(f"  {name:24s} loaded from disk (report-only)"
                  f"{'' if hist else ' — no saved history, curves not redrawn'}")
        else:
            model, hist, elapsed = train_one(name, builder, train, val,
                                             args.epochs, args.batch_size)
            (models_dir / f"{name}_history.json").write_text(
                json.dumps({k: [float(v) for v in vs] for k, vs in hist.items()}))
        trained[name] = model
        if hist is not None:
            histories[name] = hist

        p_val = model.predict(val.X, batch_size=1024, verbose=0).ravel()
        thr = pick_threshold(val.y, p_val)          # frozen from validation
        p_test = model.predict(test.X, batch_size=1024, verbose=0).ravel()
        p_train = model.predict(train.X, batch_size=1024, verbose=0).ravel()

        for split, y, p, base in (("train", train.y, p_train, baselines["train"]),
                                  ("val", val.y, p_val, baselines["val"]),
                                  ("test", test.y, p_test, baselines["test"])):
            m = classification_metrics(y, p, thr, base)
            m.update({"model": name, "split": split,
                      "params": int(model.count_params()),
                      "train_minutes": round(elapsed / 60, 2),
                      "epochs_run": len(hist["loss"]) if hist else None})
            rows.append(m)

        regimes = test_df.reset_index(drop=True).loc[test.row_index, "regime"].to_numpy()
        rb = regime_breakdown(test.y, p_test, regimes, thr)
        rb.insert(0, "model", name)
        regime_rows.append(rb)

        strat = directional_strategy(p_test, fwd_test, thr)
        strat.update({"model": name})
        strat_rows.append(strat)

        cs = cost_sensitivity(p_test, fwd_test, thr)
        cs.insert(0, "model", name)
        cs["break_even_bps"] = cs.attrs["break_even_bps"]
        cost_rows.append(cs)

        agree_rows.append({
            "model": name,
            "agrees_with_persistence": persistence_agreement(p_test, prev_ret_test, thr),
            "break_even_cost_bps": cs.attrs["break_even_bps"],
            "flips_per_session": strat["flips_per_session"],
        })

        if not args.report_only:
            model.save(models_dir / f"{name}.keras")

    summary = pd.DataFrame(rows)
    front = ["model", "split", "n", "accuracy", "baseline_persistence",
             "lift_over_persistence", "beats_persistence", "auc", "mcc", "f1",
             "precision", "recall", "brier", "threshold", "predicted_up_rate",
             "tp", "fp", "tn", "fn", "params", "epochs_run", "train_minutes"]
    summary = summary[[c for c in front if c in summary.columns]]

    np.savez(models_dir / "scaler.npz", mean=scaler.mean_, std=scaler.std_)
    (models_dir / "feature_columns.json").write_text(json.dumps(cols, indent=2))

    print("\n=== TEST RESULTS ===")
    t = summary[summary["split"] == "test"]
    for _, r in t.iterrows():
        flag = "BEATS" if r["beats_persistence"] else "below"
        print(f"  {r['model']:24s} acc={r['accuracy']*100:6.2f}%  "
              f"auc={r['auc']:.4f}  mcc={r['mcc']:+.4f}  "
              f"{flag} persistence ({r['baseline_persistence']*100:.2f}%)  "
              f"lift={r['lift_over_persistence']*100:+.2f}pp")

    preds = pd.DataFrame({"row_index": test.row_index, "y_true": test.y})
    for name, model in trained.items():
        preds[f"p_{name}"] = model.predict(test.X, batch_size=1024, verbose=0).ravel()
    preds["timestamp"] = test_df.reset_index(drop=True).loc[test.row_index, "timestamp"].to_numpy()
    preds["regime"] = test_df.reset_index(drop=True).loc[test.row_index, "regime"].to_numpy()
    preds.to_parquet(config.PROCESSED_DIR / "predictions_test.parquet", index=False)

    fig_dir = config.REPORTS_DIR / "figures"
    if histories:
        plot_curves(histories, fig_dir / "04_training_curves.png")
    plot_comparison(summary, baselines, fig_dir / "04_model_comparison.png")

    meta = {"n_features": len(cols), "n_train": len(train),
            "n_val": len(val), "n_test": len(test)}
    xlsx = config.REPORTS_DIR / "04_models.xlsx"
    with pd.ExcelWriter(xlsx, engine="openpyxl") as xw:
        pd.DataFrame({"_": []}).to_excel(xw, sheet_name="README", index=False)
        summary.to_excel(xw, sheet_name="Metrics", index=False)
        pd.concat(regime_rows, ignore_index=True).to_excel(
            xw, sheet_name="By_Regime", index=False)
        pd.DataFrame(strat_rows).to_excel(xw, sheet_name="Strategy_Check", index=False)
        pd.concat(cost_rows, ignore_index=True).to_excel(
            xw, sheet_name="Cost_Sensitivity", index=False)
        pd.DataFrame(agree_rows).to_excel(xw, sheet_name="Edge_Attribution", index=False)
        pd.DataFrame({"feature": cols}).to_excel(xw, sheet_name="Features", index=False)

        wb = xw.book
        wb._named_styles["Normal"].font = Font(name=BODY_FONT, size=10)
        ws = wb["README"]
        ws.delete_rows(1, ws.max_row)
        write_readme(ws, meta, baselines, summary)
        for sheet, w in (("Metrics", 24), ("By_Regime", 24),
                         ("Strategy_Check", 24), ("Cost_Sensitivity", 24),
                         ("Edge_Attribution", 24), ("Features", 26)):
            style(wb[sheet], wb[sheet].max_row - 1, wb[sheet].max_column, w)

    print(f"\nexcel  -> {xlsx.relative_to(config.PROJECT_ROOT)}")
    print(f"models -> models/*.keras")


if __name__ == "__main__":
    main()
